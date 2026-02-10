from openpyxl import load_workbook
from django.conf import settings

def rellenar_columna_pedido(detalles):
    """
    detalles viene así:
    [
      {'producto__nombre': 'PAPAS', 'total_producto': 120},
      {'producto__nombre': 'TOMATES', 'total_producto': 80},
    ]
    """

    ruta = settings.EXCEL_PLANTILLA_PATH
    wb = load_workbook(ruta)
    ws = wb.active

    # 1️⃣ BUSCAR EN QUÉ COLUMNA ESTÁ EL TÍTULO "PEDIDO"
    columna_pedido = None

    for fila in ws.iter_rows(min_row=1, max_row=10):  # buscamos en las primeras filas
        for celda in fila:
            if celda.value and str(celda.value).strip().upper() == "PEDIDO":
                columna_pedido = celda.column  # número de columna
                break
        if columna_pedido:
            break

    if columna_pedido is None:
        raise Exception("No se encontró la columna 'PEDIDO' en el Excel.")

    # 2️⃣ RECORRER TODO EL EXCEL BUSCANDO LOS NOMBRES DE PRODUCTOS
    for fila in ws.iter_rows(min_row=1):

        celda_producto = fila[0]  # asumimos que el nombre está en la primera columna de cada sección
        if not celda_producto.value:
            continue

        nombre_excel = str(celda_producto.value).strip().upper()

        for d in detalles:
            nombre_bd = str(d['producto__nombre']).strip().upper()
            total = d['total_producto']

            # 3️⃣ SI EL NOMBRE COINCIDE → ESCRIBIMOS EN "PEDIDO"
            if nombre_excel == nombre_bd:
                ws.cell(row=celda_producto.row, column=columna_pedido).value = float(total)
                break

    # 4️⃣ GUARDAMOS EL MISMO EXCEL
    wb.save(ruta)
