from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.views.generic import DetailView, ListView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.admin.views.decorators import staff_member_required
from django.utils import timezone
from django.db.models import Sum
from django.http import HttpResponse, JsonResponse
from django.conf import settings

from openpyxl import Workbook, load_workbook
import os
import shutil

from applications.productos.models import Producto
from .models import Pedido, DetallePedido
from applications.usuarios.models import Direccion
from django.contrib import messages
from django.urls import reverse
# =========================
# 🛒 CARRITO
# =========================

@login_required
def agregar_al_pedido(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)

    try:
        cantidad = int(request.GET.get("cantidad", 1))
    except:
        cantidad = 1

    if cantidad <= 0:
        cantidad = 1

    pedido, _ = Pedido.objects.get_or_create(
        usuario=request.user,
        estado='pendiente'
    )

    detalle, creado = DetallePedido.objects.get_or_create(
        pedido=pedido,
        producto=producto
    )

    if creado:
        detalle.cantidad = cantidad
    else:
        detalle.cantidad += cantidad

    detalle.save()
    return JsonResponse({"ok": True})


class PedidoPendienteDetailView(LoginRequiredMixin, DetailView):
    model = Pedido
    template_name = 'pedidos/ver_pedido.html'
    context_object_name = 'pedido'

    def get_object(self):
        return Pedido.objects.filter(
            usuario=self.request.user,
            estado='pendiente'
        ).first()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pedido = self.object
        detalles = pedido.detallepedido_set.all() if pedido else []
        context['detalles'] = detalles
        context['total_cantidad'] = sum(d.cantidad for d in detalles)
        return context


@login_required
@require_POST
def actualizar_cantidad(request, detalle_id):
    detalle = get_object_or_404(
        DetallePedido,
        id=detalle_id,
        pedido__usuario=request.user
    )

    try:
        cantidad = int(request.POST.get('cantidad', detalle.cantidad))
    except:
        cantidad = detalle.cantidad

    if cantidad > 0:
        detalle.cantidad = cantidad
        detalle.save()
    else:
        detalle.delete()

    return redirect('ver_pedido')


@login_required
@require_POST
def eliminar_item(request, detalle_id):
    detalle = get_object_or_404(
        DetallePedido,
        id=detalle_id,
        pedido__usuario=request.user
    )
    detalle.delete()
    return redirect('ver_pedido')


class ConfirmarPedidoView(LoginRequiredMixin, View):
    def post(self, request):
        pedido = Pedido.objects.filter(usuario=request.user, estado='pendiente').first()

        if not pedido or not pedido.detallepedido_set.exists():
            messages.info(request, "No tienes productos en el pedido.")
            return redirect('ver_pedido')

        direccion_id = request.POST.get('direccion_id')
        direccion = None

        if direccion_id:
            direccion = get_object_or_404(Direccion, pk=direccion_id, usuario=request.user)

        # Si no seleccionó, usar principal
        if not direccion:
            direccion = request.user.direcciones.filter(principal=True).first()

        # Si aún no hay dirección → enviar a direcciones
        if not direccion:
            messages.error(request, "Debes crear o seleccionar una dirección antes de confirmar.")
            return redirect(f"{reverse('lista_direcciones')}?next=pedido")

        # ✅ Confirmar pedido
        pedido.direccion = direccion
        pedido.estado = 'confirmado'
        pedido.save()

        messages.success(request, "✅ Pedido confirmado correctamente.")
        return redirect('mis_pedidos')


class MisPedidosListView(LoginRequiredMixin, ListView):
    model = Pedido
    template_name = 'pedidos/mis_pedidos.html'
    context_object_name = 'pedidos'

    def get_queryset(self):
        return Pedido.objects.filter(
            usuario=self.request.user
        ).order_by('-fecha')


class PedidoDetailView(DetailView):
    model = Pedido
    template_name = "pedidos/ver_pedido.html"
    context_object_name = "pedido"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        detalles = DetallePedido.objects.filter(pedido=self.object)
        context["detalles"] = detalles
        context["total_cantidad"] = sum(item.cantidad for item in detalles)
        return context


# =========================
# ✅ RESUMEN DIARIO
# =========================

@staff_member_required
def resumen_diario(request):
    hoy = timezone.now().date()

    pedidos = Pedido.objects.filter(
        estado='confirmado',
        fecha__date=hoy
    )

    detalles = DetallePedido.objects.filter(
        pedido__in=pedidos
    ).values(
        'producto__nombre'
    ).annotate(
        total_producto=Sum('cantidad')
    ).order_by('producto__nombre')

    total_kilos = sum(d['total_producto'] for d in detalles)
    total_pedidos = pedidos.count()

    return render(request, 'pedidos/resumen_diario.html', {
        'detalles': detalles,
        'total_kilos': total_kilos,
        'total_pedidos': total_pedidos,
        'hoy': hoy,
    })


# =========================
# ✅ EXCEL RESUMEN SIMPLE
# =========================

@staff_member_required
def exportar_resumen_excel(request):
    hoy = timezone.now().date()

    pedidos = Pedido.objects.filter(
        estado='confirmado',
        fecha__date=hoy
    )

    detalles = DetallePedido.objects.filter(
        pedido__in=pedidos
    ).values(
        'producto__nombre'
    ).annotate(
        total_producto=Sum('cantidad')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Diario"

    ws.append(["Producto", "Total Kilos"])

    total_kilos = 0
    for d in detalles:
        ws.append([d['producto__nombre'], d['total_producto']])
        total_kilos += d['total_producto']

    ws.append([])
    ws.append(["TOTAL KILOS", total_kilos])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=resumen_{hoy}.xlsx"

    wb.save(response)
    return response


# =========================
# ✅ EXCEL REAL CON LIMPIEZA DIARIA
# =========================

@staff_member_required
def generar_excel_diario(request):
    hoy = timezone.now().date()

    # ✅ PEDIDOS DEL DÍA
    detalles = DetallePedido.objects.filter(
        pedido__fecha__date=hoy,
        pedido__estado='confirmado'
    )

    resumen = {}
    for item in detalles:
        nombre = item.producto.nombre.upper().strip()
        resumen[nombre] = resumen.get(nombre, 0) + item.cantidad

    # ✅ RUTA EXCEL BASE
    ruta_base = os.path.join(settings.BASE_DIR, 'reportes', 'INVENTARIO DIARIO.xlsx')

    # ✅ COPIA CON FECHA
    nombre_archivo = f"Inventario_{hoy}.xlsx"
    ruta_copia = os.path.join(settings.BASE_DIR, 'reportes', nombre_archivo)

    shutil.copy(ruta_base, ruta_copia)

    # ✅ ESCRIBIR EN COPIA
    wb = load_workbook(ruta_copia)
    ws = wb.active

    for fila in ws.iter_rows(min_row=1):

        # FRUTAS (A → D)
        fruta = str(fila[0].value).upper().strip() if fila[0].value else ""
        if fruta in resumen:
            fila[3].value = resumen[fruta]

        # VERDURAS (H → K)
        verdura = str(fila[7].value).upper().strip() if fila[7].value else ""
        if verdura in resumen:
            fila[10].value = resumen[verdura]

    wb.save(ruta_copia)

    # ✅ LIMPIAR EXCEL BASE
    wb_base = load_workbook(ruta_base)
    ws_base = wb_base.active

    for fila in ws_base.iter_rows(min_row=1):
        fila[3].value = ""
        fila[10].value = ""

    wb_base.save(ruta_base)

    # ✅ DESCARGA AUTOMÁTICA
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'

    with open(ruta_copia, "rb") as f:
        response.write(f.read())

    return response
